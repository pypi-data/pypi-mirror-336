from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO, TextIOWrapper
import csv


class FileCreator:
    def __init__(self, headers):
        self.headers = headers
        self.rows = []
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self._initialize_excel_headers()
        self._add_data_validation()
        self.in_memory_file = BytesIO()

    def _initialize_excel_headers(self):
        """Initialize the Excel sheet with the provided headers."""
        for col_num, header in enumerate(self.headers, start=1):
            self.sheet.cell(row=1, column=col_num, value=header)

    def _add_data_validation(self):
        """Tambah dropdown Valid/Tidak Valid pada kolom Status Verifikasi."""
        try:
            col_index = self.headers.index("Status Verifikasi") + 1  # Cari indeks kolom
            col_letter = get_column_letter(col_index)

            # Buat Data Validation untuk dropdown Valid/Tidak Valid
            dv = DataValidation(type="list", formula1='"Valid,Tidak Valid"', showDropDown=False)

            self.sheet.add_data_validation(dv)

            # Terapkan validasi ke semua baris di kolom Status Verifikasi (misalnya hingga 1000 baris)
            dv.add(f"{col_letter}2:{col_letter}1000")

            print(f"Data validation berhasil ditambahkan pada kolom {col_letter}")
        except ValueError:
            print("Kolom 'Status Verifikasi' tidak ditemukan dalam header.")

    def add_row(self, row_data):
        """Add a row to both the Excel sheet and the internal CSV rows list."""
        if len(row_data) != len(self.headers):
            raise ValueError("Row data must match the number of headers.")

        # Append the row to the Excel sheet
        self.sheet.append(row_data)

        # Append the row to the internal CSV list
        self.rows.append(row_data)

    def hide_columns(self, *columns):
        """
        Hide specified columns in the Excel sheet.
        Columns can be provided as 1-based indices (e.g., 1, 2, 3) or names (e.g., "A", "B").
        """
        for col in columns:
            if isinstance(col, int):  # If column is given as an index
                col_letter = get_column_letter(col)
            else:
                col_letter = col  # Assume it's already a column letter
            self.sheet.column_dimensions[col_letter].hidden = True

    def hide_rows(self, *rows):
        """
        Hide specified rows in the Excel sheet.
        Rows should be provided as 1-based indices.
        """
        for row in rows:
            self.sheet.row_dimensions[row].hidden = True

    def save_excel(self):
        """Save the Excel file to an in-memory BytesIO object."""
        self.workbook.save(self.in_memory_file)
        self.in_memory_file.seek(0)  # Reset the file pointer to the beginning of the file

    def save_csv(self):
        """Save the CSV file to an in-memory BytesIO object."""
        self.in_memory_file = BytesIO()  # Reset the in-memory file

        # Use TextIOWrapper to wrap the BytesIO object for text (str) writing
        text_wrapper = TextIOWrapper(self.in_memory_file, encoding='utf-8', newline='')

        csv_writer = csv.writer(text_wrapper, quoting=csv.QUOTE_MINIMAL)

        # Write headers and rows to the CSV file
        csv_writer.writerow(self.headers)
        csv_writer.writerows(self.rows)

        # Flush the wrapper and seek the file back to the start
        text_wrapper.flush()
        self.in_memory_file.seek(0)
        text_wrapper.detach()  # Detach the wrapper to avoid closing the BytesIO object

    def get_file(self):
        """Return the in-memory file."""
        return self.in_memory_file