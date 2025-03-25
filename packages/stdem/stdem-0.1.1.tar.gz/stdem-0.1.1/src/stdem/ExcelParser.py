import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

import json

from . import HeadType

class Head:

    def __init__(self, sheet: Worksheet, row: tuple[Cell, ...]) -> None:
        self.sheet = sheet
        self.column = len(row)
        self.head = HeadType.headCreater(row[0])
        self.headList: list[HeadType.HeadType] = [self.head] * self.column

    def getCellMaxCol(self, cell: Cell) -> int:
        for i in self.sheet.merged_cells.ranges:
            if cell.coordinate in i:
                return i.max_col - 1
        return cell.column - 1

    def rowParser(self, row: tuple[Cell, ...]) -> None:
        i = 0
        while i < self.column:
            if row[i].value:
                h = HeadType.headCreater(row[i])
                j = self.getCellMaxCol(row[i])
                self.headList[i].addChild(h)
                self.headList[i:j] = [h] * (j - i)
                i = j
            else:
                i += 1



def getData(filename: str) -> HeadType.data:
    workbook = openpyxl.load_workbook(filename)

    iter_rows = workbook.active.iter_rows()
    head = Head(workbook.active, next(iter_rows)[1:])

    isData = False
    for row in iter_rows:
        if row[0].value == "#":
            continue
        elif row[0].value == "#data":
            isData = True
            dataRoot = head.head.parsetData(row[1:], True)
            continue

        if isData:
            head.head.parsetData(row[1:], False)
        else:
            head.rowParser(row[1:])
    return dataRoot


def getJson(filename: str) -> str:
    return json.dumps(getData(filename))


if __name__ == "__main__":
    print(getJson("Table.xlsx"))
