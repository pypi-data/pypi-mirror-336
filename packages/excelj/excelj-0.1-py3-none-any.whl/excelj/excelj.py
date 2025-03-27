import openpyxl # type: ignore
import json
import os
import argparse


class Eparse:
    def __init__(self,excel_path:str):
        self.workbook = openpyxl.load_workbook(filename=excel_path)
    @staticmethod
    def remove_index(tup, indexs):
        return tuple(x for i, x in filter(lambda ix: ix[0] not in indexs, enumerate(tup)))
    def sheet_filter(self,sheet_name:str=None,start_line_index:int=1)->tuple[list,str]:
    
        sheetnames = self.workbook.sheetnames
        sheet_name = sheetnames[0] if sheet_name==None  else sheet_name
        sh = self.workbook[sheet_name]
        first_line =  list(sh.values)[start_line_index]
        null_indexs = []
        for i in range(len(first_line)):
            if i==0:
                continue
            if first_line[i] == None:
                null_indexs.append(i)
        filtered = []
        rows = list(sh.values)[start_line_index:]
        for row in rows: 
                if row[0] == None:
                    continue
                filtered.append(self.remove_index(row, null_indexs))


        return (filtered,sheet_name)
    def sheet_to_json(self,sheet_name:str=None,start_line_index:int=0)->tuple[str,str]:
        result={}
        data = self.sheet_filter(sheet_name,start_line_index)
        rows = data[0]
        sheet_name = data[1]
        print(f"正在解析  {sheet_name}  ...")
        first_row = rows[0]
        for row_index in range(len(rows)):
            if row_index==0:
                continue
            row = rows[row_index]
            result[str(row[0])]={
                first_row[i]:row[i] for i in range(len(row)) if i!=0
            }
        return (json.dumps(result,ensure_ascii=False,sort_keys=True, indent=4, separators=(',', ':')),sheet_name)

    def sheet_to_json_auto(self,sheet_names:str=[],start_line_index:int=0)->list[tuple[str,str]]:
        result:list[tuple[str,str]]=[]
        for sheet_name in sheet_names:
           result.append(self.sheet_to_json(sheet_name,start_line_index))
        return result
    
    def sheet_to_json_all(self,start_line_index:int=0)->list[tuple[str,str]]:
        return self.sheet_to_json_auto(self.workbook.sheetnames,start_line_index)


def parse_args():
    parser = argparse.ArgumentParser(
        description="帮助命令")
    parser.add_argument('p', help="需要解析的excel文件的路径")
    parser.add_argument('-s', default=os.getcwd(), help="解析后保存的路径,默认保存在当前目录")
    parser.add_argument('-r', default=0, type=int, help="从第几行开始解析,默认从第一行开始解析")
    parser.add_argument('-a', default=True, type=int, help="是否自动解析所有sheet,默认为真")
    args = parser.parse_args()
    return args

def save_to_file(data:tuple[str,str],savePath:str=os.getcwd()):
    save_name = data[1]
    save_data = data[0]
    save_path = fr"{savePath}\{save_name}.json"
    with open(save_path, "w",encoding='utf-8') as f:
        f.write(save_data)
    print("已保存到:  "+save_path)



def main():
    args=parse_args()
    excelPath = args.p
    savePath = args.s
    start_index = args.r
    is_parse_all = args.a
    ep = Eparse(excelPath)
    if is_parse_all:
        for i in ep.sheet_to_json_all(start_line_index=start_index):
            save_to_file(i,savePath)
    else:

        save_to_file(ep.sheet_to_json(start_line_index=start_index,savePath=savePath))
            



r'''
示例：
python jsonparse.py "C:\Users\zhang\Desktop\test.xlsx" -s "C:\Users\zhang\Desktop" -r 1 -a False


'''

