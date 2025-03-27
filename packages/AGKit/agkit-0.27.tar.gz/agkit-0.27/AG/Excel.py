import os

import xlrd
from openpyxl.reader.excel import load_workbook

from AG.Logger import Logger


class Excel:
    @staticmethod
    def read_excel(file_path):
        if not os.path.exists(file_path):
            Logger.Instance.Error("文件不存在！")
            return {}

        data = {}
        _, ext = os.path.splitext(file_path)

        try:
            if ext.lower() == '.xls':
                workbook = xlrd.open_workbook(file_path)
                for sheet in workbook.sheets():
                    for row_idx in range(sheet.nrows):
                        row = sheet.row(row_idx)
                        cells = [cell.value for cell in row]
                        data[row_idx] = cells
            elif ext.lower() in ('.xlsx', '.xlsm'):
                workbook = load_workbook(file_path, read_only=True)
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_idx = sheet.iter_rows().index(row)
                        data[row_idx] = list(row)
            else:
                Logger.Instance.Error("不支持的文件格式")
                return {}
        except Exception as e:
            Logger.Instance.Error(f"读取文件失败: {str(e)}")
            return {}
        return data

    @staticmethod
    def GetSheets(path):
        """
        获取工作簿
        :param path:excel路径
        :return:excel中的工作簿信息
        """
        if Excel.IsExcelValid(path):
            excel = xlrd.open_workbook(path)
            return excel.sheets()
        return None

    @staticmethod
    def IsExcelValid(path):
        if path.endswith(".xls"):
            return True
        Logger.Instance.Error("%s文件格式异常，请确保表格文件为.xls格式" % path)
        return False

    @staticmethod
    def IsSheetValid(sheet):
        """
        工作簿是否有效
        :param sheet:工作簿信息
        :return:是否有效
        """
        info = sheet.name.split("_")
        if info[0].lower() == "no":
            return False
        return True

    @staticmethod
    def GetCellValue(sheet, row, col):
        """
        获取工作簿单元格内容
        :param sheet:工作簿
        :param row:行
        :param col:列
        :return:单元格值
        """
        cell = sheet.cell(row, col)
        c_type = cell.ctype
        c_value = cell.value
        if c_type == 0:
            return ""
        elif c_type == 1:
            return c_value
        elif c_type == 2:
            if c_value % 1 == 0.0:
                return int(c_value)
            else:
                return c_value
