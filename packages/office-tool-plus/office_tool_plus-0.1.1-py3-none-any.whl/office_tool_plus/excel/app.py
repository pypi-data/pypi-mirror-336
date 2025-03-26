from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from ..utils import *


class ExcelTools:
    def __init__(self):
        # 获取当前系统平台信息
        self.platform = platform.system()
        self.app = None

    def create_app(self):
        from win32com.client import gencache
        if self.app is None:
            self.app = gencache.EnsureDispatch("Excel.Application")
            self.app.Visible = False

    def close_app(self):
        self.app.Quit()
        self.app = None

    @check_platform('windows')
    def single_to_pdf(self, excel_path, sheet_names=None, pdf_dir=None):
        """
        将指定的Excel工作簿导出为PDF格式。

        参数:
        - excel_path: str, Excel文件的路径。
        - sheet_names: list, 需要导出的工作表名称列表。如果未提供，则默认导出所有工作表。
        - pdf_dir: str, PDF文件的保存目录。如果未提供，则默认保存在Excel文件的同目录下。

        返回:
        - pdf_path: 导出的PDF文件路径。
        """

        # 检查并准备文件路径
        excel_path = check_file_path(excel_path)
        pdf_path = save_file_path(excel_path, ".pdf", pdf_dir)

        # 如果没有提供特定的工作表名称，则导出所有工作表
        if self.app is None:
            self.create_app()

        # 打开Excel工作簿
        workbook = self.app.Workbooks.Open(excel_path)

        # 确定需要导出的工作表名称列表
        sheet_names = sheet_names or [sheet.Name for sheet in workbook.Sheets]

        # 识别出需要隐藏的工作表
        sheets_to_hide = [sheet for sheet in workbook.Sheets if sheet.Name not in sheet_names]
        # 临时隐藏不需要导出的工作表
        for sheet in sheets_to_hide:
            sheet.Visible = False

        try:
            # 导出可见的工作表为PDF
            workbook.ExportAsFixedFormat(0, str(pdf_path))
        except Exception as e:
            print(f"导出Excel到PDF时出错：{e}")

        # 恢复所有工作表的可见性
        for sheet in sheets_to_hide:
            sheet.Visible = True
        # 关闭时不保存更改
        workbook.Close(SaveChanges=False)
        # 退出应用程序
        self.app.Quit()
        self.app = None
        # 返回导出的PDF文件路径
        return pdf_path

    @check_platform('windows')
    def many_to_pdf(self, excel_dir, suffix=None, recursive=True, pdf_dir=None):
        """
        将指定目录下的Excel文件转换为PDF格式。

        参数:
        - excel_dir: str,  Excel文件所在目录
        - pdf_dir: str, PDF文件的保存目录。如果未提供，则默认保存在Excel文件的同目录下。
        """
        suffix = suffix or ["*.xlsx", "*.xls"]
        excel_path_yield = search_files(excel_dir, suffix, recursive)
        if self.app is None:
            self.create_app()
        for excel_path in excel_path_yield:
            # 检查并准备文件路径
            pdf_path = save_file_path(excel_path, ".pdf", pdf_dir)
            # 打开Excel工作簿
            workbook = self.app.Workbooks.Open(excel_path)
            try:
                # 导出可见的工作表为PDF
                workbook.ExportAsFixedFormat(0, str(pdf_path))
            except Exception as e:
                print(f"导出工作表到PDF时出错：{e}")
            # 关闭工作簿时不保存更改
            workbook.Close(SaveChanges=False)

        # 退出Excel应用程序
        self.app.Quit()
        self.app = None

    def from_template(self, template_file, labor_datas, sheet_name=None, output_dir=None):
        """
        根据模板文件和模板数据生成新的文件。

        :param template_file: 模板文件的路径。
        :param labor_datas: 包含模板数据的列表，每个模板数据包括文件名和要写入的数据。
        :param sheet_name: 指定模板工作表，默认为第一个工作表。
        :param output_dir: 输出文件的目录，默认为模板文件所在目录。
        """
        # 确保模板文件路径有效
        abs_template_file = check_file_path(template_file)
        # 获取路径、文件名和后缀
        output_dir = output_dir or os.path.dirname(abs_template_file)
        template_name = os.path.basename(abs_template_file)
        tmp_file_name, tmp_file_suffix = os.path.splitext(template_name)

        # 遍历每个模板数据项
        for lb in labor_datas:
            # 打开模板文件
            tmp_wb = load_workbook(abs_template_file)
            # 选择指定的工作表或默认活动工作表
            tmp_ws = sheet_name or tmp_wb.active
            # 获取新文件名，默认使用模板文件名
            new_file_name = lb.get("文件名", tmp_file_name)
            # 获取要写入的数据
            cell_data = lb.get("数据", [])
            # 遍历每个数据项并写入工作表
            for cell_value in cell_data:
                for k, v in cell_value.items():
                    # 根据数据类型插入图像或文本
                    if v["type"] == "image":
                        self.insert_image(tmp_ws, v["value"], k)
                    else:
                        tmp_ws[k] = v["value"]
            # 生成唯一的输出文件路径
            output_file = unique_file_path(str(os.path.join(output_dir, new_file_name + tmp_file_suffix)))
            # 保存修改后的文件
            tmp_wb.save(output_file)
            # 关闭工作簿释放资源
            tmp_wb.close()

    @staticmethod
    def insert_image(ws, img_path, cell_address):
        """
        将图片插入到指定的工作表单元格中，并根据单元格的大小调整图片尺寸。

        参数:
            ws (openpyxl.worksheet.worksheet.Worksheet): 目标工作表对象。
            img_path (str): 图片文件的路径。
            cell_address (str): 图片插入的目标单元格地址（例如 'A1'）。

        返回值:
            无返回值。该函数直接修改工作表对象，将图片插入到指定位置。
        """

        # 加载图片
        img = Image(img_path)

        # 获取目标单元格对象
        cell = ws[cell_address]

        # 假设DPI为96，用于将行高从点转换为像素
        dpi = 96

        # 检查目标单元格是否为合并单元格
        merged_cells = [merged for merged in ws.merged_cells.ranges if cell.coordinate in merged]
        if merged_cells:
            # 如果是合并单元格，获取合并范围的起始和结束行列号
            merged_range = merged_cells[0]
            start_col, start_row, end_col, end_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )

            # 计算合并单元格的总宽度（以像素为单位）
            width = sum(
                (ws.column_dimensions[get_column_letter(col)].width or 8) * 7  # 列宽默认值为8
                for col in range(start_col, end_col + 1)
            )

            # 计算合并单元格的总高度（以像素为单位）
            height = sum(
                (ws.row_dimensions[row].height or 15) * dpi / 72  # 行高默认值为15
                for row in range(start_row, end_row + 1)
            )
        else:
            # 如果不是合并单元格，直接计算单元格的宽度和高度
            width = (ws.column_dimensions[cell.column_letter].width or 8) * 7  # 列宽默认值为8
            height = (ws.row_dimensions[cell.row].height or 15) * dpi / 72  # 行高默认值为15

        # 根据单元格的宽度和高度调整图片的尺寸
        img.width = width
        img.height = height

        # 设置图片的锚点位置为目标单元格
        img.anchor = cell_address

        # 将调整后的图片添加到工作表中
        ws.add_image(img)
