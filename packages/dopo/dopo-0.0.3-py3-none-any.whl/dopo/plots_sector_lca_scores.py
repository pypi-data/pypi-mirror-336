"""
This module processes Excel files to create dot plots and stacked bar charts using OpenPyXL. It 
categorizes sheets by sector and generates charts for each sector, including scatter plots for LCA 
scores and stacked bar charts for data contributions. Customizations include axis labels, titles, 
and visual styles.
"""

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, BarChart, Reference, Series

def _categorize_sheets_by_sector(file_path):
    """
    Categorizes the sheets in an Excel workbook by sector.

    This function reads an Excel workbook and categorizes the sheets based on 
    the sector, assuming that the sector name is the first part of the sheet 
    name separated by an underscore ('_'). Sheets without an underscore in their 
    name are skipped.

    Parameters
    ----------
    file_path : str
        Path to the Excel workbook file.

    Returns
    -------
    dict
        A dictionary where the keys are sector names and the values are lists of 
        sheet names corresponding to that sector.
    """
    # Load the workbook
    workbook = load_workbook(filename=file_path, read_only=True)
    
    # Initialize a dictionary to hold sectors and their corresponding sheet names
    worksheet_dict = {}
    
    # Iterate over all sheet names in the workbook
    for sheet_name in workbook.sheetnames:
        # Skip combined sector sheets (assuming these sheets don't have an underscore)
        if '_' not in sheet_name:
            continue
        
        # Split the sheet name to extract the sector (assumes sector is the first part)
        sector = sheet_name.split('_')[0]
        
        # Add the sheet name to the corresponding sector in the dictionary
        if sector in worksheet_dict:
            worksheet_dict[sector].append(sheet_name)
        else:
            worksheet_dict[sector] = [sheet_name]
    
    return worksheet_dict

def dot_plots_xcl(filepath_workbook, column_positions):
    """
    Creates dot plots for each sector in an Excel workbook.

    This function reads an Excel workbook, categorizes sheets by sector, and creates 
    scatter charts (dot plots) for data visualization based on the input data in the 
    worksheets. It saves the generated charts in a new sheet within the workbook.

    Parameters
    ----------
    filepath_workbook : str
        Path to the Excel workbook file.
    index_positions : dict
        A dictionary containing column index positions for the data required to create 
        the charts for each worksheet.

    Returns
    -------
    int
        The row position where the last chart was placed.
    """
    worksheet_dict = _categorize_sheets_by_sector(filepath_workbook)
    
    # Load the workbook
    wb = load_workbook(filepath_workbook)
    
    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)        
                
        # Initial position for the first chart
        current_row = 1  # Start placing charts from row 1
        current_col = 1  # Start placing charts from column 1
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        charts_per_row = 3  # Number of charts per row
        
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # Find min_row, max_row and max_column
            max_row = ws.max_row
            max_column = ws.max_column
            min_row = 1

            # Find the key in index_positions that contains worksheet_name
            matching_key = None
            for key in column_positions.keys():
                if worksheet_name in key:
                    matching_key = key
                    break

            if not matching_key:
                print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
                continue

            # Retrieve the column positions from the index_positions dictionary
            positions = column_positions[matching_key]
            total_col = positions.get("total", None) + 1
            rank_col = positions.get("rank", None) + 1
            mean_col = positions.get("mean", None) + 1
            std_adv_col = positions.get("2std_abv", None) + 1
            std_blw_col = positions.get("2std_blw", None) + 1
            q1_col = positions.get("q1", None) + 1
            q3_col = positions.get("q3", None) + 1
            method_col = positions.get("method", None) + 1
            method_unit_col = positions.get("method unit", None) + 1
            
            # Ensure that all required columns are present
            if None in [total_col, rank_col, mean_col, std_adv_col, std_blw_col, q1_col, q3_col, 
                        method_col, method_unit_col]:
                print(f"Warning: Missing columns in worksheet '{worksheet_name}' for sector '{sector}'. Skipping...")
                continue
            
            # Create a ScatterChart (or other chart type as needed)
            chart = ScatterChart()

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            chart.title = f"{method_value} LCA scores for {sector} sector" 
            
            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.y_axis.title = f"{method_unit_value}"
            chart.x_axis.title = 'activity rank'
            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 

            # Define the data range for the chart
            y_values = Reference(ws, min_col=total_col, min_row=min_row, max_row=max_row)
            x_values = Reference(ws, min_col=rank_col, min_row=min_row, max_row=max_row)

            # Create a series and add it to the chart
            series = Series(y_values, x_values, title_from_data=True)
            chart.series.append(series)
            chart.style = 9

            # Customize the series to show only markers (dots)
            series.marker.symbol = "circle"
            series.marker.size = 5
            series.graphicalProperties.line.noFill = True

            # Adjust X-axis properties
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.majorGridlines = None 
            chart.x_axis.tickMarkSkip = 1  # Show all tick marks
            chart.x_axis.tickLblSkip = 1  # Show all labels

            chart.x_axis.scaling.orientation = "minMax"
            chart.x_axis.crosses = "autoZero"
            chart.x_axis.axPos = "b"
            chart.x_axis.delete = False

            # Adjust Y-axis properties
            chart.y_axis.tickLblPos = "nextTo"  # Position the labels next to the tick marks
            chart.y_axis.delete = False  # Ensure axis is not deleted
            chart.y_axis.number_format = '0.00000'
            chart.y_axis.majorGridlines = None 

            # Add statistics: mean, IQR, and standard deviation lines to the chart
            # MEAN
            mean_y = Reference(ws, min_col=mean_col, min_row=min_row, max_row=max_row)
            mean_series = Series(mean_y, x_values, title_from_data="True")
            chart.series.append(mean_series)
            mean_series.marker.symbol = "none"  # No markers, just a line
            mean_series.graphicalProperties.line.solidFill = "FF0000"  # Red line for mean value
            mean_series.graphicalProperties.line.width = 10000  # Set line width

            # IQR
            iqr1 = Reference(ws, min_col=q1_col, min_row=min_row, max_row=max_row)
            iqr3 = Reference(ws, min_col=q3_col, min_row=min_row, max_row=max_row)
            iqr1_series = Series(iqr1, x_values, title_from_data="True")
            iqr3_series = Series(iqr3, x_values, title_from_data="True")
            chart.series.append(iqr1_series)
            chart.series.append(iqr3_series)
            iqr1_series.marker.symbol = "none"  # No markers, just a line
            iqr3_series.marker.symbol = "none"
            iqr1_series.graphicalProperties.line.solidFill = "6082B6"  # Blue line 
            iqr3_series.graphicalProperties.line.solidFill = "6082B6"  
            iqr1_series.graphicalProperties.line.width = 10000  # Set line width
            iqr3_series.graphicalProperties.line.width = 10000  # Set line width

            # STD
            std_abv = Reference(ws, min_col=std_adv_col, min_row=min_row, max_row=max_row)
            std_blw = Reference(ws, min_col=std_blw_col, min_row=min_row, max_row=max_row)
            std_abv_series = Series(std_abv, x_values, title_from_data="True")
            std_blw_series = Series(std_blw, x_values, title_from_data="True")
            chart.series.append(std_abv_series)
            chart.series.append(std_blw_series)
            std_abv_series.marker.symbol = "none"  # No markers, just a line
            std_blw_series.marker.symbol = "none"
            std_abv_series.graphicalProperties.line.solidFill = "FFAA1D"  # Orange line
            std_blw_series.graphicalProperties.line.solidFill = "FFAA1D"  
            std_abv_series.graphicalProperties.line.width = 10000  # Set line width
            std_blw_series.graphicalProperties.line.width = 10000  # Set line width

            # Set legend position to the right of the plot area
            chart.legend.position = 'r'  # 'r' for right
            chart.legend.overlay = False

            # Adjust chart dimensions
            chart.width = 20  # Width of the chart
            chart.height = 14  # Height of the chart

            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)
            
            # Update position for the next chart
            current_col += chart_width +1 
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height +1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)

    wb.save(filepath_workbook)
    return current_row

def stacked_bars_xcl(filepath_workbook, column_positions, current_row_dot_plot):
    """
    Creates stacked bar charts for each sector in an Excel workbook.

    This function reads an Excel workbook, categorizes sheets by sector, and creates 
    stacked bar charts to visualize data contributions. The generated charts are added 
    to a new or existing sheet within the workbook.

    Parameters
    ----------
    filepath_workbook : str
        Path to the Excel workbook file.
    index_positions : dict
        A dictionary containing column index positions for the data required to create 
        the charts for each worksheet.
    current_row_dot_plot : int
        The row number in the chart sheet where the dot plots ended, used to determine 
        the starting row for the stacked bar charts.

    Returns
    -------
    int
        The row position where the last chart was placed.
    """
    # Categorize sheets by sector
    worksheet_dict = _categorize_sheets_by_sector(filepath_workbook)
    
    # Load the workbook
    wb = load_workbook(filepath_workbook)
    
    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)
                
        # Initial position for the first chart
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        current_row = current_row_dot_plot + chart_height  # Start placing charts from row after dot plots
        current_col = 1  # Start placing charts from column 1
        charts_per_row = 3  # Number of charts per row
        
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # Find the key in index_positions that contains worksheet_name
            matching_key = None
            for key in column_positions.keys():
                if worksheet_name in key:
                    matching_key = key
                    break

            if not matching_key:
                print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
                continue

            # Retrieve the column positions from the index_positions dictionary
            positions = column_positions[matching_key]

            # Find min_row, max_row, and max_column
            max_row = ws.max_row
            max_column = ws.max_column
            input_min_col = positions.get("first_input", None) + 1
            rank_col = positions.get("rank", None) + 1
            method_col = positions.get("method", None) + 1
            method_unit_col = positions.get("method unit", None) + 1

            # Create a BarChart object for the stacked bar chart
            chart = BarChart()
            chart.type = "bar"
            chart.style = 2
            chart.grouping = "stacked"
            chart.overlap = 100

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            chart.title = f"Inputs contributions to {method_value} LCA score for sector {sector}"

            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.y_axis.title = f"{method_unit_value}"
            chart.x_axis.title = 'activity rank'

            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 
            chart.legend.overlay = False

            # Define data for the stacked bar chart
            data = Reference(ws, min_col=input_min_col, min_row=1, max_row=max_row, 
                             max_col=max_column)
            cats = Reference(ws, min_col=rank_col, min_row=2, max_row=max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            # Modify each series in the chart to disable the inversion of negative values 
            for series in chart.series:
                series.invertIfNegative = False

            # y-axis ticks
            chart.y_axis.tickLblPos = "nextTo"
            chart.y_axis.delete = False  # Ensure axis is not deleted
            chart.y_axis.number_format = '0.000'

            # Adjust X-axis properties
            chart.x_axis.tickLblPos = "low" 
            chart.x_axis.majorGridlines = None 
            chart.x_axis.tickMarkSkip = 1  # Show all tick marks
            chart.x_axis.tickLblSkip = 1  # Show all labels

            chart.x_axis.scaling.orientation = "minMax"
            chart.x_axis.crosses = "autoZero"
            chart.x_axis.axPos = "b"
            chart.x_axis.delete = False

            # Adjust chart dimensions
            chart.width = 20  # Width of the chart
            chart.height = 14  # Height of the chart

            # Add the chart to the chart worksheet
            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)
            
            # Update position for the next chart
            current_col += chart_width + 1
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height + 1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)
        
    # Save the workbook with the charts added
    wb.save(filepath_workbook)
    
    return current_row
