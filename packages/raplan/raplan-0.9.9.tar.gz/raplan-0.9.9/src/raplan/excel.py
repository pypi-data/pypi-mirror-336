"""RaPlan I/O with Excel. For this functionality to work, the 'excel' extra should be
installed.
"""

import json
import uuid
from collections.abc import Callable
from dataclasses import field as field_init
from pathlib import Path
from typing import Any, get_args

import serde
import serde.json

from raplan import distributions
from raplan.classes import Component, Horizon, Maintenance, Project, System, Task

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.styles.protection import Protection
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError(
        "Package 'openpyxl' not found, please install it yourself "
        "or via the 'excel' package extra of RaPlan."
    )

_TYPES_TO_SHEETS = {
    Project: "Project",
    System: "Systems",
    Component: "Components",
    Maintenance: "Maintenance",
    Task: "Tasks",
    Horizon: "Horizon",
}
_SHEETS_TO_TYPES = {v: k for k, v in _TYPES_TO_SHEETS.items()}
_FIELDS = dict()
for title, cls in _SHEETS_TO_TYPES.items():
    _FIELDS[title] = list(getattr(cls, "__dataclass_fields__", dict()).keys())
    _FIELDS[title].remove("uuid")  # Remove and add uuid as first column.
    _FIELDS[title].insert(0, "uuid")


def to_excel(
    project: Project, path: str | Path | None = None, lock: bool = False
) -> openpyxl.Workbook:
    """Convert a RaPlan project to an Excel workbook.

    Arguments:
        project: RaPlan project to export.
        path: Optional filename to save the workbook to.

    Returns:
        An openpyxl Workbook.
    """
    wb = Workbook()

    # Let schedule be the first sheet.
    schedule = wb.active
    schedule.title = "Schedule"
    schedule.append(
        [
            "project",
            "system",
            "component",
            "maintenance",
            "task",
            "maintenance_time",
            "task_rejuvenation",
            "task_duration",
            "task_cost",
        ]
    )
    if lock:
        _lock_worksheet(schedule)

    # Create sheets for instance storage.
    sheets: dict[str, Worksheet] = dict()
    indices: dict[str, dict[uuid.UUID, int]] = dict()
    objects: dict[uuid.UUID, Any] = dict()

    for cls, title in _TYPES_TO_SHEETS.items():  # Setup sheets, fields and headers.
        sheets[title] = wb.create_sheet(title)
        sheets[title].append(_FIELDS[title])
        indices[title] = dict()
        if lock:
            _lock_worksheet(sheets[title])

    # Add data.
    horizon = project.horizon
    _add_obj_row(project, sheets, indices, objects, horizon)
    _add_obj_row(project.horizon, sheets, indices, objects, horizon)
    for system in project.systems:
        _add_obj_row(system, sheets, indices, objects, horizon)
        for component in system.components:
            _add_obj_row(component, sheets, indices, objects, horizon)
            for maintenance in component.maintenance:
                _add_obj_row(maintenance, sheets, indices, objects, horizon)
                _add_obj_row(maintenance.task, sheets, indices, objects, horizon)
                _add_schedule_row(
                    schedule,
                    project,
                    system,
                    component,
                    maintenance,
                    wb,
                    indices,
                )

    _set_defined_names(wb)  # Add local defined names for UUIDs.

    for ws in wb.worksheets:  # Set style and lock all worksheets.
        _set_table_style(ws)
        _set_auto_width(ws)

    if path:
        wb.save(str(Path(path)))

    return wb


def _add_obj_row(
    obj: Any,
    sheets: dict[str, Worksheet],
    indices: dict[str, dict[uuid.UUID, int]],
    objects: dict[uuid.UUID, Any],
    horizon: Horizon,
):
    key = _TYPES_TO_SHEETS[type(obj)]
    ws, fields = sheets[key], _FIELDS[key]
    if obj.uuid in objects:
        return
    row = [_get_cell_value(obj.uuid, sheets)]
    hyperlinks = []
    for field_index, field in enumerate(fields[1:]):
        value = getattr(obj, field, None)

        if "time" in field and value is not None and type(value) in {int, float}:
            value += horizon.start

        if isinstance(value, list):
            for iter_index, entry in enumerate(value):
                _ensure_header(ws, field, iter_index)
                row.append(_get_cell_value(entry, sheets))
                if type(entry) in _TYPES_TO_SHEETS:
                    # field_index + 2 because of for loop offset
                    hyperlinks.append(field_index + iter_index + 2)
            continue

        if type(value) in _TYPES_TO_SHEETS:
            hyperlinks.append(field_index + 2)

        row.append(_get_cell_value(value, sheets))

    ws.append(row)

    row_index = ws.max_row
    for col in hyperlinks:  # Make hyperlinks to defined names.
        _make_hyperlink(ws.cell(row_index, col))

    indices[key][obj.uuid] = row_index  # Add UUID to index list.
    objects[obj.uuid] = obj  # Add object to reference dict.


def _add_schedule_row(
    ws: Worksheet,
    project: Project,
    system: System,
    component: Component,
    maintenance: Maintenance,
    wb: Workbook,
    indices: dict[str, dict[uuid.UUID, int]],
):
    """
    project,
    system,
    component,
    maintenance,
    task,
    maintenance_time,
    task_rejuvenation,
    task_duration,
    task_cost,
    """
    ref_objs: list[Project | System | Component | Maintenance | Task] = [
        project,
        system,
        component,
        maintenance,
        maintenance.task,
    ]
    ref_uuids = [x.uuid for x in ref_objs]

    # Start with plain references
    row = [_encode_uuid(x) for x in ref_uuids]
    ws.append(row)
    for cell, obj in zip(ws[ws.max_row][: len(ref_objs)], ref_objs):
        cell.hyperlink = Hyperlink(
            ref="",
            location=cell.value,
        )
        if obj.name is not None:
            cell.value = obj.name

    objs: list[Project | System | Component | Maintenance | Task] = [
        maintenance,
        maintenance.task,
        maintenance.task,
        maintenance.task,
    ]
    props = [
        "time",
        "rejuvenation",
        "duration",
        "cost",
    ]
    src_row = ws.max_row
    src_col = len(row) + 1
    for obj, prop in zip(objs, props):
        _move_obj_value(obj, prop, ws.cell(src_row, src_col), wb, indices)
        src_col += 1


def _get_cell_value(value: Any, sheets: dict[str, Worksheet]) -> Any:
    if isinstance(value, uuid.UUID):
        return _encode_uuid(value)
    if isinstance(value, get_args(distributions.Distributions)):
        return serde.json.to_json({value.__class__.__name__: value})
    t = type(value)
    if t in _TYPES_TO_SHEETS:
        return _encode_uuid(value.uuid)
    return value


def _encode_uuid(value: uuid.UUID) -> str:
    return "_" + str(value).replace("-", "_")


def _ensure_header(ws: Worksheet, field: str, index: int):
    if index == 0:
        return
    columns = [c.value for c in ws[1]]
    name_index = f"{field}-{index}"
    if name_index in columns:
        return
    index_0 = columns.index(field) + 1
    ws.insert_cols(index_0 + 1)
    ws.cell(1, index_0 + 1, name_index)


def _move_obj_value(
    obj,
    field: str,
    target: Cell,
    wb: Workbook,
    indices: dict[str, dict[uuid.UUID, int]],
):
    key = _TYPES_TO_SHEETS[type(obj)]
    src_row = indices[key][obj.uuid]
    src_col = _FIELDS[key].index(field) + 1
    source = wb[key].cell(src_row, src_col)
    _move_value(wb, source, target, f"{key}_{field}{_encode_uuid(obj.uuid)}")


def _move_value(wb: Workbook, source: Cell, target: Cell, name: str):
    ref = f"{target.parent.title}!${target.column_letter}${target.row}"
    target.value = source.value
    try:
        _add_defined_name(wb, name, ref)
        _make_hyperlink(source, name)
        source.value = f"={ref}"
        _make_editable(target)
    except ValueError as e:
        # A task or maintenance object may have been re-used.
        if "already exists" in str(e):
            target.value = f"={ref}"
            _make_hyperlink(target, name)
        else:
            raise e


def _make_hyperlink(cell: Cell, name: str | None = None):
    location = name or str(cell.value)
    cell.hyperlink = Hyperlink(ref="", location=location)


def _make_editable(cell: Cell):
    cell.protection = Protection(locked=False)


def _add_defined_name(wb: Workbook, name: str, ref: str):
    wb.defined_names[name] = openpyxl.workbook.defined_name.DefinedName(
        name,
        attr_text=ref,
    )


def _set_defined_names(wb: Workbook):
    for ws in wb.worksheets[1:]:
        for row in ws[f"A2:A{ws.max_row}"]:
            cell = row[0]
            column_letter = get_column_letter(ws.max_column)
            _add_defined_name(
                wb,
                cell.value,
                f"{ws.title}!$B${cell.row}:${column_letter}${cell.row}",
            )
        ws.column_dimensions["A"].hidden = True


def _set_table_style(ws: Worksheet):
    if ws.max_row < 2:  # openpyxl can't make a proper table without content.
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    name = ws.title
    tab = Table(name=name, displayName=name, ref=ref)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def _set_auto_width(ws: Worksheet):
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].bestFit = True


def _lock_worksheet(ws: Worksheet):
    ws.protection.sheet = True
    ws.protection.password = "raplan"
    ws.protection.enable()


def from_excel(path: str | Path) -> Project:
    """Convert an Excel file to a RaPlan project.

    Arguments:
        path: Path to an Excel workbook containing a RaPlan project.
    """
    wb = openpyxl.load_workbook(path)
    objects: dict[uuid.UUID, Any] = dict()

    # First set horizon
    horizon = None

    for sheetname in [
        "Horizon",
        "Tasks",
        "Maintenance",
        "Components",
        "Systems",
    ]:
        ws = wb[sheetname]
        for i in range(2, ws.max_row + 1):
            _from_row(ws[i], sheetname, wb, objects, horizon)

        if sheetname == "Horizon":
            horizon = [item for item in objects.values()][0]
            horizon.start = int(horizon.start)
            horizon.end = int(horizon.end)

    ws = wb["Project"]
    project = _from_row(ws[2], "Project", wb, objects)
    return project


def _decode_uuid(enc: str):
    return uuid.UUID(enc[1:].replace("_", "-"), version=4)


def _decode_distribution(enc: str):
    d = json.loads(enc)
    for k, v in d.items():
        return serde.from_dict(getattr(distributions, k), v)


_CELL_VALUE_CONVERSION: dict[Any, Callable[[Any], Any]] = {
    int | float: float,
    int | float | None: float,
    str | None: str,
    str: str,
    uuid.UUID: _decode_uuid,
    distributions.Distributions: _decode_distribution,
}


def _from_row(
    row: tuple[Cell, ...],
    sheetname: str,
    wb: Workbook,
    objects: dict[uuid.UUID, Any],
    horizon: Horizon = None,
):
    cls = _SHEETS_TO_TYPES[sheetname]
    fields = _FIELDS[sheetname]
    kwargs = dict()
    for field in fields:
        value = _arg_from_row(row, field, sheetname, wb, objects)
        if value is None:
            continue

        if "time" in field and type(value) in {int, float} and horizon:
            value -= horizon.start

        kwargs[field] = value
    obj = cls(**kwargs)
    objects[obj.uuid] = obj
    return obj


def _arg_from_row(
    row: tuple[Cell, ...],
    field: str,
    sheetname: str,
    wb: Workbook,
    objects: dict[uuid.UUID, Any],
):
    cls = _SHEETS_TO_TYPES[sheetname]

    t = getattr(getattr(cls, "__dataclass_fields__", dict())[field], "type", None)
    index = _FIELDS[sheetname].index(field)
    value = _resolve_value(wb, row[index])
    if value == "" or value is None or t is None:
        return None

    if t in _CELL_VALUE_CONVERSION:
        return _CELL_VALUE_CONVERSION[t](value)

    if t in _TYPES_TO_SHEETS:
        return objects[_decode_uuid(value)]

    field_spec = cls.__dataclass_fields__.get(field, field_init())
    if (field_spec.default_factory,) == (list,):  # Ignores linters asking for isinstance(...)
        headers = [c.value for c in wb[sheetname][1]]
        result = [objects[_decode_uuid(value)]]
        offset = 1
        while f"{field}-{offset}" in headers:
            enc_uuid = _resolve_value(wb, row[index + offset])
            if enc_uuid is None or enc_uuid == "":
                break
            result.append(objects[_decode_uuid(enc_uuid)])
            offset += 1
        return result

    return None


def _resolve_value(wb: Workbook, cell: Cell) -> str:
    if str(cell.value).startswith("="):
        parts = str(cell.value)[1:].split("!")
        if len(parts) < 2:
            return _resolve_value(wb, cell.parent[parts[0]])
        return _resolve_value(wb, wb[parts[0]][parts[1]])
    else:
        return cell.value
